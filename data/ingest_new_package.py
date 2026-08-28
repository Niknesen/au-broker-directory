#!/usr/bin/env python3
"""
One-off ingestion of the new, deduped broker package (16,771 unique
businesses, matched by Google Place ID - a hard dedup key the previous
~12,148-row dataset never had, which is why it turned out to be only ~5,620
truly unique records under the hood).

Replaces data/all_brokers_full.json - this new file supersedes the old one
for anything sourced from the Google Places master spreadsheet (merging
would risk reintroducing the duplicates this package was built to remove).

The one thing it DOES carry over from the old file: the 545 rows with an
"outreach-*" id, a separate one-off batch merged in earlier from an AIEX
outreach list (brokersemailoutreach.xlsx) that was never part of the
Places-sourced master spreadsheet at all, so this package can't regenerate
it. 335 of those 545 turned out to already exist in the new package (same
business, matched by phone) and are skipped; the other 210 - including
ACY Securities - only exist in that old outreach batch and are appended
back in with fresh "outreach-N" ids.

Also writes data/reviews_by_place.json: real Google review text (rating,
text, reviewer, relative time), grouped by Place ID, for the broker detail
pages to render. The first package this was built from had a scraper bug
where every review's text was literally the string "[object Object]" - this
run is against the corrected re-export where 96% of rows have real text.

What this does NOT do, and why:
  - Does not carry over the old `about` text. The old dataset had no stable
    join key (no Place ID) to match old rows to new ones reliably, and only
    a minority of old rows had `about` populated anyway. Every row here
    gets about=None, same fallback behaviour the site already handles.
  - Drops CLOSED_PERMANENTLY / CLOSED_TEMPORARILY rows and the ~59 rows
    (0.35%) whose Address doesn't parse as a real AU address (a few are
    literally overseas places from a bad geocode match).
"""
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
SOURCE_XLSX = Path("/Users/nick/Downloads/final-australian-broker-package (1).xlsx")
OUT = ROOT / "all_brokers_full.json"
REVIEWS_OUT = ROOT / "reviews_by_place.json"
MAX_REVIEWS_PER_BUSINESS = 12
MAX_REVIEW_CHARS = 700

ADDR_RE = re.compile(r"([A-Za-z0-9 '\.-]+?)\s+(NSW|VIC|QLD|WA|SA|TAS|ACT|NT)\s+(\d{4})")

# Our 7-bucket taxonomy the site already uses (category chips, page copy).
# Rows from the "original" 5,620 already carry one of these exact names in
# the Industry column. Rows from the "expansion" 11,151 only carry raw
# Google Places types (Category column, pipe-delimited) - map those.
BUCKETS = [
    "Mortgage & Finance",
    "Insurance",
    "Real Estate & Buyers",
    "Business Sales & Franchise",
    "Asset & Equipment Finance",
    "Customs & Freight",
    "Wealth & Investment",
]

FRIENDLY_TYPE = {
    "real_estate_agency": "Real estate agency",
    "insurance_agency": "Insurance agency",
    "accounting": "Accounting firm",
    "finance": "Finance broker",
    "consultant": "Finance broker",
}


def bucket_from_google_types(raw_category):
    tokens = {t.strip() for t in (raw_category or "").split("|")}
    if "real_estate_agency" in tokens:
        return "Real Estate & Buyers", "Real estate agency"
    if "insurance_agency" in tokens:
        return "Insurance", "Insurance agency"
    if "accounting" in tokens:
        return "Wealth & Investment", "Accounting firm"
    return "Mortgage & Finance", "Finance broker"


def parse_address(address):
    m = ADDR_RE.search(address or "")
    if not m:
        return None
    suburb, state, postcode = m.groups()
    suburb = suburb.split(",")[-1].strip(" ,")
    return suburb, state, postcode


def first_email(raw):
    if not raw:
        return None
    return raw.split(";")[0].strip() or None


def main():
    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    ws = wb["Brokers"]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    idx = {name: i for i, name in enumerate(header)}

    dropped_closed = 0
    dropped_bad_address = 0
    bucket_counts = Counter()
    out = []

    for row in it:
        status = row[idx["Business Status"]]
        if status in ("CLOSED_PERMANENTLY", "CLOSED_TEMPORARILY"):
            dropped_closed += 1
            continue

        address = row[idx["Address"]]
        parsed = parse_address(address)
        if not parsed:
            dropped_bad_address += 1
            continue
        suburb, state, postcode = parsed

        industry = row[idx["Industry"]]
        raw_category = row[idx["Category"]]
        if industry in BUCKETS:
            sheet = industry
            category_label = raw_category or industry
        else:
            sheet, category_label = bucket_from_google_types(raw_category)

        bucket_counts[sheet] += 1
        n = bucket_counts[sheet]

        out.append({
            "id": f"{sheet}-{n}",
            "sheet": sheet,
            "name": row[idx["Business"]],
            "category": category_label,
            "phone": row[idx["Phone"]],
            "email": first_email(row[idx["Public Email"]]),
            "website": row[idx["Website URL"]],
            "address": address,
            "suburb": suburb,
            "city": suburb,
            "state": state,
            "postcode": postcode,
            "rating": row[idx["Rating"]],
            "reviews": row[idx["Reviews"]] or 0,
            "maps_url": row[idx["Google Maps URL"]],
            "place_id": row[idx["Place ID"]],
            "about": None,
            "about_source": None,
            "team": [],
            "license_disclosed": False,
            "license_detail": None,
        })

    added_outreach = merge_outreach_batch(out, bucket_counts)

    with open(OUT, "w") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    print(f"Wrote {len(out)} brokers to {OUT} ({added_outreach} from the outreach batch)")
    print(f"Dropped: {dropped_closed} closed, {dropped_bad_address} bad address")
    for name, count in bucket_counts.most_common():
        print(f"  {name}: {count}")

    ingest_reviews(wb)


def norm_phone(p):
    return re.sub(r"\D", "", p or "")[-9:]


def merge_outreach_batch(out, bucket_counts):
    """Bring back the ~210 businesses that only ever existed in the old
    one-off outreach merge (see module docstring) - skip the ones that
    turned out to already be in the new package under the same phone
    number."""
    old_path = ROOT / "all_brokers_full.json.bak-pre-16k"
    if not old_path.exists():
        print("No pre-16k backup found - skipping outreach-batch merge.")
        return 0

    with open(old_path) as f:
        old = json.load(f)
    outreach = [b for b in old if str(b.get("id", "")).startswith("outreach-")]

    existing_phones = {norm_phone(b.get("phone")) for b in out if b.get("phone")}
    added = 0
    for b in outreach:
        phone_key = norm_phone(b.get("phone"))
        if phone_key and phone_key in existing_phones:
            continue
        sheet = b.get("sheet") or "Mortgage & Finance"
        bucket_counts[sheet] += 1
        out.append({
            "id": f"outreach-{added + 1}",
            "sheet": sheet,
            "name": b["name"],
            "category": b.get("category") or sheet,
            "phone": b.get("phone"),
            "email": b.get("email"),
            "website": b.get("website"),
            "address": b.get("address"),
            "suburb": b.get("suburb"),
            "city": b.get("city"),
            "state": b.get("state"),
            "postcode": b.get("postcode"),
            "rating": b.get("rating"),
            "reviews": b.get("reviews") or 0,
            "maps_url": b.get("maps_url"),
            "place_id": None,
            "about": None,
            "about_source": None,
            "team": [],
            "license_disclosed": False,
            "license_detail": None,
        })
        if phone_key:
            existing_phones.add(phone_key)
        added += 1
    return added


def ingest_reviews(wb):
    ws = wb["Reviews"]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    idx = {name: i for i, name in enumerate(header)}

    by_place = {}
    total = 0
    skipped_no_text = 0
    for row in it:
        total += 1
        text = (row[idx["Review Text"]] or "").strip()
        if not text:
            skipped_no_text += 1
            continue
        if len(text) > MAX_REVIEW_CHARS:
            text = text[:MAX_REVIEW_CHARS].rsplit(" ", 1)[0] + "…"
        place_id = row[idx["Place ID"]]
        bucket = by_place.setdefault(place_id, [])
        if len(bucket) >= MAX_REVIEWS_PER_BUSINESS:
            continue
        bucket.append({
            "rating": row[idx["Review Rating"]],
            "text": text,
            "reviewer": row[idx["Reviewer"]] or "Google user",
            "relative_time": row[idx["Relative Time"]],
        })

    with open(REVIEWS_OUT, "w") as f:
        json.dump(by_place, f, indent=2, ensure_ascii=False)

    print(f"\nWrote reviews for {len(by_place)} businesses to {REVIEWS_OUT}")
    print(f"Reviews processed: {total}, skipped (no text): {skipped_no_text}")


if __name__ == "__main__":
    main()
